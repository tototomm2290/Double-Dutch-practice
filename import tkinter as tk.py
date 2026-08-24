import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook, load_workbook
import streamlit as st
import pandas as pd  # 表データをWeb上で動的に扱うために使用



class ExcelLikeApp:

    def __init__(self, root):

        self.root = root
        self.root.title("○× 試行管理表")
        self.root.geometry("1400x800")

        # -----------------------------------------
        # 初期設定
        # -----------------------------------------
        self.trial_count = 10
        self.item_count = 10

        # セルを保存
        self.cells = {}

        # -----------------------------------------
        # 上部メニュー
        # -----------------------------------------
        top = tk.Frame(root)
        top.pack(fill="x", padx=5, pady=5)

        tk.Label(
            top,
            text="○× 試行管理表",
            font=("Meiryo", 16, "bold")
        ).pack(side="left", padx=10)

        tk.Button(
            top,
            text="Excelを開く",
            command=self.open_excel
        ).pack(side="right", padx=5)

        tk.Button(
            top,
            text="Excel保存",
            command=self.save_excel
        ).pack(side="right", padx=5)

        tk.Button(
            top,
            text="列を追加",
            command=self.add_trial_column
        ).pack(side="right", padx=5)

        tk.Button(
            top,
            text="行を追加",
            command=self.add_item
        ).pack(side="right", padx=5)

        # -----------------------------------------
        # スクロール
        # -----------------------------------------
        outer = tk.Frame(root)
        outer.pack(fill="both", expand=True)

        self.canvas = tk.Canvas(
            outer,
            bg="white"
        )
        self.canvas.pack(
            side="left",
            fill="both",
            expand=True
        )

        self.v_scroll = ttk.Scrollbar(
            outer,
            orient="vertical",
            command=self.canvas.yview
        )
        self.v_scroll.pack(
            side="right",
            fill="y"
        )

        self.h_scroll = ttk.Scrollbar(
            root,
            orient="horizontal",
            command=self.canvas.xview
        )
        self.h_scroll.pack(
            side="bottom",
            fill="x"
        )

        self.canvas.configure(
            yscrollcommand=self.v_scroll.set,
            xscrollcommand=self.h_scroll.set
        )

        # -----------------------------------------
        # 表
        # -----------------------------------------
        self.table_frame = tk.Frame(
            self.canvas,
            bg="white"
        )

        self.canvas.create_window(
            (0, 0),
            window=self.table_frame,
            anchor="nw"
        )

        self.table_frame.bind(
            "<Configure>",
            self.update_scroll_region
        )

        # -----------------------------------------
        # 表を作成
        # -----------------------------------------
        self.create_table()

        # Ctrl + S
        self.root.bind(
            "<Control-s>",
            lambda event: self.save_excel()
        )

    # =====================================================
    # スクロール範囲
    # =====================================================

    def update_scroll_region(self, event=None):

        self.canvas.configure(
            scrollregion=self.canvas.bbox("all")
        )

    # =====================================================
    # セルを登録
    # =====================================================

    def register_cell(self, row, col, widget):

        self.cells[(row, col)] = widget

        widget.grid(
            row=row,
            column=col,
            padx=1,
            pady=1,
            sticky="nsew"
        )

    # =====================================================
    # 表作成
    # =====================================================

    def create_table(self):

        # A1
        blank = tk.Label(
            self.table_frame,
            text="",
            bg="#eeeeee",
            relief="solid",
            borderwidth=1,
            width=14,
            height=2
        )

        self.register_cell(
            0,
            0,
            blank
        )

        # 試行番号
        for col in range(1, self.trial_count + 1):

            header = tk.Label(
                self.table_frame,
                text=str(col),
                bg="#d9eaf7",
                relief="solid",
                borderwidth=1,
                font=("Meiryo", 10, "bold"),
                width=10,
                height=2
            )

            self.register_cell(
                0,
                col,
                header
            )

        # ○割合
        percentage_header = tk.Label(
            self.table_frame,
            text="○割合",
            bg="#ffe699",
            relief="solid",
            borderwidth=1,
            font=("Meiryo", 10, "bold"),
            width=12,
            height=2
        )

        self.register_cell(
            0,
            self.trial_count + 1,
            percentage_header
        )

        # 項目
        for i in range(self.item_count):

            choice_row = 1 + i * 2
            memo_row = choice_row + 1

            self.create_item(
                i,
                choice_row,
                memo_row
            )

    # =====================================================
    # 項目を1つ作成
    # =====================================================

    def create_item(
        self,
        item_index,
        choice_row,
        memo_row
    ):

        # ---------------------------------
        # 項目名
        # ---------------------------------

        item = tk.Entry(
            self.table_frame,
            font=("Meiryo", 10),
            relief="solid",
            borderwidth=1,
            width=18
        )

        item.insert(
            0,
            f"項目{item_index + 1}"
        )

        item.bind(
            "<KeyRelease>",
            lambda event, w=item:
            self.text_resize(w)
        )

        # 最後の項目なら自動追加
        item.bind(
            "<FocusIn>",
            lambda event, r=choice_row:
            self.auto_add_item_if_last(r)
        )

        self.register_cell(
            choice_row,
            0,
            item
        )

        # ---------------------------------
        # メモ
        # ---------------------------------

        memo_label = tk.Label(
            self.table_frame,
            text="メモ",
            bg="#f3f3f3",
            relief="solid",
            borderwidth=1,
            width=18
        )

        self.register_cell(
            memo_row,
            0,
            memo_label
        )

        # ---------------------------------
        # ○×セル
        # ---------------------------------

        for col in range(
            1,
            self.trial_count + 1
        ):

            combo = ttk.Combobox(
                self.table_frame,
                values=["", "○", "×"],
                state="readonly",
                justify="center",
                font=("Meiryo", 12),
                width=8
            )

            combo.set("")

            combo.bind(
                "<<ComboboxSelected>>",
                lambda event,
                r=choice_row,
                c=col,
                w=combo:
                self.choice_changed(
                    r,
                    c,
                    w
                )
            )

            # 最後の列なら自動追加
            combo.bind(
                "<FocusIn>",
                lambda event, c=col:
                self.auto_add_column_if_last(c)
            )

            self.register_cell(
                choice_row,
                col,
                combo
            )

            # ---------------------------------
            # メモ
            # ---------------------------------

            memo = tk.Entry(
                self.table_frame,
                font=("Meiryo", 9),
                relief="solid",
                borderwidth=1,
                width=10
            )

            memo.bind(
                "<KeyRelease>",
                lambda event, w=memo:
                self.text_resize(w)
            )

            self.register_cell(
                memo_row,
                col,
                memo
            )

        # ---------------------------------
        # ○割合
        # ---------------------------------

        percentage = tk.Label(
            self.table_frame,
            text="0.0%",
            bg="#fff2cc",
            relief="solid",
            borderwidth=1,
            font=("Meiryo", 10, "bold"),
            width=12
        )

        self.register_cell(
            choice_row,
            self.trial_count + 1,
            percentage
        )

        # メモ側
        memo_percentage = tk.Label(
            self.table_frame,
            text="",
            bg="#fff8df",
            relief="solid",
            borderwidth=1,
            width=12
        )

        self.register_cell(
            memo_row,
            self.trial_count + 1,
            memo_percentage
        )

    # =====================================================
    # ○ / × の色変更
    # =====================================================

    def choice_changed(
        self,
        row,
        col,
        widget
    ):

        value = widget.get()

        # ○
        if value == "○":

            widget.configure(
                foreground="green",
                background="#d9ead3"
            )

        # ×
        elif value == "×":

            widget.configure(
                foreground="red",
                background="#f4cccc"
            )

        # 空欄
        else:

            widget.configure(
                foreground="black"
            )

        # 割合を更新
        self.update_percentage(row)

        # 最終列なら自動追加
        if col == self.trial_count:

            self.add_trial_column()

    # =====================================================
    # ○割合計算
    # =====================================================

    def update_percentage(self, row):

        count_o = 0
        count_total = 0

        for col in range(
            1,
            self.trial_count + 1
        ):

            widget = self.cells.get(
                (row, col)
            )

            if widget is None:
                continue

            value = widget.get()

            if value in ["○", "×"]:

                count_total += 1

            if value == "○":

                count_o += 1

        # 1つも入力されていない場合
        if count_total == 0:

            percentage = 0

        else:

            percentage = (
                count_o /
                count_total
            ) * 100

        percentage_widget = self.cells.get(
            (
                row,
                self.trial_count + 1
            )
        )

        if percentage_widget:

            percentage_widget.config(
                text=f"{percentage:.1f}%"
            )

    # =====================================================
    # 文字数に応じてサイズ変更
    # =====================================================

    def text_resize(self, widget):

        try:
            text = widget.get()

        except Exception:
            return

        length = len(text)

        # 横幅
        width = 10 + length // 2

        if width < 10:
            width = 10

        if width > 40:
            width = 40

        widget.config(
            width=width
        )

        # 長文なら高さを増やす
        row = widget.grid_info()["row"]

        if length > 25:

            self.table_frame.grid_rowconfigure(
                row,
                minsize=50
            )

        elif length > 50:

            self.table_frame.grid_rowconfigure(
                row,
                minsize=70
            )

        self.update_scroll_region()

    # =====================================================
    # 列を追加
    # =====================================================

    def add_trial_column(self):

        old_trial_count = self.trial_count

        new_col = old_trial_count + 1

        # ---------------------------------
        # 古い割合列を退避
        # ---------------------------------

        old_percentage_col = old_trial_count + 1

        old_header = self.cells.get(
            (0, old_percentage_col)
        )

        if old_header:

            old_header.grid_forget()

        for i in range(self.item_count):

            choice_row = 1 + i * 2
            memo_row = choice_row + 1

            p1 = self.cells.get(
                (choice_row, old_percentage_col)
            )

            p2 = self.cells.get(
                (memo_row, old_percentage_col)
            )

            if p1:
                p1.grid_forget()

            if p2:
                p2.grid_forget()

        # ---------------------------------
        # 試行回数更新
        # ---------------------------------

        self.trial_count += 1

        # ---------------------------------
        # 新しい番号
        # ---------------------------------

        header = tk.Label(
            self.table_frame,
            text=str(new_col),
            bg="#d9eaf7",
            relief="solid",
            borderwidth=1,
            font=("Meiryo", 10, "bold"),
            width=10,
            height=2
        )

        self.register_cell(
            0,
            new_col,
            header
        )

        # ---------------------------------
        # ○× + メモ
        # ---------------------------------

        for i in range(self.item_count):

            choice_row = 1 + i * 2
            memo_row = choice_row + 1

            combo = ttk.Combobox(
                self.table_frame,
                values=["", "○", "×"],
                state="readonly",
                justify="center",
                font=("Meiryo", 12),
                width=8
            )

            combo.bind(
                "<<ComboboxSelected>>",
                lambda event,
                r=choice_row,
                c=new_col,
                w=combo:
                self.choice_changed(
                    r,
                    c,
                    w
                )
            )

            self.register_cell(
                choice_row,
                new_col,
                combo
            )

            memo = tk.Entry(
                self.table_frame,
                font=("Meiryo", 9),
                relief="solid",
                borderwidth=1,
                width=10
            )

            memo.bind(
                "<KeyRelease>",
                lambda event, w=memo:
                self.text_resize(w)
            )

            self.register_cell(
                memo_row,
                new_col,
                memo
            )

        # ---------------------------------
        # 新しい割合列
        # ---------------------------------

        percentage_col = self.trial_count + 1

        header = tk.Label(
            self.table_frame,
            text="○割合",
            bg="#ffe699",
            relief="solid",
            borderwidth=1,
            font=("Meiryo", 10, "bold"),
            width=12,
            height=2
        )

        self.register_cell(
            0,
            percentage_col,
            header
        )

        for i in range(self.item_count):

            choice_row = 1 + i * 2
            memo_row = choice_row + 1

            percentage = tk.Label(
                self.table_frame,
                text="0.0%",
                bg="#fff2cc",
                relief="solid",
                borderwidth=1,
                font=("Meiryo", 10, "bold"),
                width=12
            )

            self.register_cell(
                choice_row,
                percentage_col,
                percentage
            )

            memo_percentage = tk.Label(
                self.table_frame,
                text="",
                bg="#fff8df",
                relief="solid",
                borderwidth=1,
                width=12
            )

            self.register_cell(
                memo_row,
                percentage_col,
                memo_percentage
            )

            self.update_percentage(choice_row)

        self.update_scroll_region()

    # =====================================================
    # 最後の列を使ったら自動追加
    # =====================================================

    def auto_add_column_if_last(self, col):

        if col == self.trial_count:

            self.add_trial_column()

    # =====================================================
    # 行を追加
    # =====================================================

    def add_item(self):

        item = self.item_count

        choice_row = 1 + item * 2
        memo_row = choice_row + 1

        self.create_item(
            item,
            choice_row,
            memo_row
        )

        self.item_count += 1

        self.update_scroll_region()

    # =====================================================
    # 最後の行なら自動追加
    # =====================================================

    def auto_add_item_if_last(self, row):

        last_row = (
            1 +
            (self.item_count - 1) * 2
        )

        if row == last_row:

            self.add_item()

    # =====================================================
    # Excel保存
    # =====================================================

    def save_excel(self):

        filepath = filedialog.asksaveasfilename(
            title="Excelファイルとして保存",
            defaultextension=".xlsx",
            filetypes=[
                ("Excelファイル", "*.xlsx")
            ]
        )

        if not filepath:
            return

        try:

            wb = Workbook()
            ws = wb.active
            ws.title = "試行管理表"

            # ---------------------------------
            # ヘッダー
            # ---------------------------------

            ws.cell(
                row=1,
                column=1,
                value=""
            )

            for col in range(
                1,
                self.trial_count + 1
            ):

                ws.cell(
                    row=1,
                    column=col + 1,
                    value=col
                )

            ws.cell(
                row=1,
                column=self.trial_count + 2,
                value="○割合"
            )

            # ---------------------------------
            # データ
            # ---------------------------------

            for i in range(self.item_count):

                choice_row = 1 + i * 2
                memo_row = choice_row + 1

                excel_choice_row = choice_row + 1
                excel_memo_row = memo_row + 1

                # 項目名
                item_widget = self.cells.get(
                    (choice_row, 0)
                )

                if item_widget:

                    ws.cell(
                        row=excel_choice_row,
                        column=1,
                        value=item_widget.get()
                    )

                # メモ
                ws.cell(
                    row=excel_memo_row,
                    column=1,
                    value="メモ"
                )

                # 各試行
                for col in range(
                    1,
                    self.trial_count + 1
                ):

                    choice_widget = self.cells.get(
                        (choice_row, col)
                    )

                    memo_widget = self.cells.get(
                        (memo_row, col)
                    )

                    if choice_widget:

                        ws.cell(
                            row=excel_choice_row,
                            column=col + 1,
                            value=choice_widget.get()
                        )

                    if memo_widget:

                        ws.cell(
                            row=excel_memo_row,
                            column=col + 1,
                            value=memo_widget.get()
                        )

                # 割合
                percentage_widget = self.cells.get(
                    (
                        choice_row,
                        self.trial_count + 1
                    )
                )

                if percentage_widget:

                    ws.cell(
                        row=excel_choice_row,
                        column=self.trial_count + 2,
                        value=percentage_widget.cget("text")
                    )

            # ---------------------------------
            # 列幅
            # ---------------------------------

            ws.column_dimensions["A"].width = 20

            for col in range(
                2,
                self.trial_count + 2
            ):

                letter = self.number_to_excel_column(
                    col
                )

                ws.column_dimensions[
                    letter
                ].width = 12

            ws.column_dimensions[
                self.number_to_excel_column(
                    self.trial_count + 2
                )
            ].width = 12

            # ---------------------------------
            # 保存
            # ---------------------------------

            wb.save(filepath)

            messagebox.showinfo(
                "保存完了",
                "Excelファイルを保存しました。"
            )

        except Exception as e:

            messagebox.showerror(
                "保存エラー",
                f"保存できませんでした。\n\n{e}"
            )

    # =====================================================
    # Excel読み込み
    # =====================================================

    def open_excel(self):

        filepath = filedialog.askopenfilename(
            title="Excelファイルを開く",
            filetypes=[
                ("Excelファイル", "*.xlsx")
            ]
        )

        if not filepath:
            return

        try:

            wb = load_workbook(
                filepath,
                data_only=True
            )

            ws = wb.active

            # ---------------------------------
            # 最大列
            # ---------------------------------

            max_col = ws.max_column

            # ○割合の列を除く
            if max_col >= 2:

                self.trial_count = max(
                    1,
                    max_col - 2
                )

            # ---------------------------------
            # 項目数
            # ---------------------------------

            max_row = ws.max_row

            self.item_count = max(
                1,
                (max_row - 1) // 2
            )

            # ---------------------------------
            # 現在の表を削除
            # ---------------------------------

            for widget in self.table_frame.winfo_children():

                widget.destroy()

            self.cells.clear()

            # ---------------------------------
            # 新しく表を作る
            # ---------------------------------

            self.create_table()

            # ---------------------------------
            # データを読み込む
            # ---------------------------------

            for i in range(self.item_count):

                choice_row = 1 + i * 2
                memo_row = choice_row + 1

                excel_choice_row = choice_row + 1
                excel_memo_row = memo_row + 1

                # 項目名
                item_widget = self.cells.get(
                    (choice_row, 0)
                )

                if item_widget:

                    value = ws.cell(
                        row=excel_choice_row,
                        column=1
                    ).value

                    if value is not None:

                        item_widget.delete(
                            0,
                            tk.END
                        )

                        item_widget.insert(
                            0,
                            str(value)
                        )

                # 各試行
                for col in range(
                    1,
                    self.trial_count + 1
                ):

                    choice_widget = self.cells.get(
                        (choice_row, col)
                    )

                    memo_widget = self.cells.get(
                        (memo_row, col)
                    )

                    # ○ / ×
                    if choice_widget:

                        value = ws.cell(
                            row=excel_choice_row,
                            column=col + 1
                        ).value

                        if value is not None:

                            choice_widget.set(
                                str(value)
                            )

                            self.set_choice_color(
                                choice_widget
                            )

                    # メモ
                    if memo_widget:

                        value = ws.cell(
                            row=excel_memo_row,
                            column=col + 1
                        ).value

                        if value is not None:

                            memo_widget.insert(
                                0,
                                str(value)
                            )

                self.update_percentage(
                    choice_row
                )

            self.update_scroll_region()

            messagebox.showinfo(
                "読み込み完了",
                "Excelファイルを読み込みました。"
            )

        except Exception as e:

            messagebox.showerror(
                "読み込みエラー",
                f"Excelファイルを読み込めませんでした。\n\n{e}"
            )

    # =====================================================
    # ○×の色
    # =====================================================

    def set_choice_color(self, widget):

        value = widget.get()

        if value == "○":

            widget.configure(
                foreground="green"
            )

        elif value == "×":

            widget.configure(
                foreground="red"
            )

        else:

            widget.configure(
                foreground="black"
            )

    # =====================================================
    # Excel列番号
    # =====================================================

    def number_to_excel_column(self, number):

        result = ""

        while number > 0:

            number, remainder = divmod(
                number - 1,
                26
            )

            result = chr(
                65 + remainder
            ) + result

        return result

# ---- 2. データの初期化（ここだけに1回書く） ----
if "my_data" not in st.session_state:
    st.session_state.my_data = load_data()

# ブラウザ上に直接編集できる表を表示（文字数調整やセル直接編集に自動対応）
edited_data = st.data_editor(
    st.session_state.my_data,
    num_rows="dynamic"  # 行・列の自動追加を許可
)

# 保存ボタンが押された時の処理
if st.button("Excelに保存"):
    # 既存のopenpyxl等を使った保存関数を呼び出す
    既存の保存処理関数(edited_data)
    st.success("保存が完了しました！")
    


# =========================================================
# メイン
# =========================================================

if __name__ == "__main__":

    root = tk.Tk()

    app = ExcelLikeApp(root)

    root.mainloop()
    
