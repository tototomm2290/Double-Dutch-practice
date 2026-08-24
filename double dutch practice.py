import io
import streamlit as st
import pandas as pd
from openpyxl import Workbook, load_workbook

# =========================================================
# ページ設定
# =========================================================

st.set_page_config(
    page_title="○×試行管理表",
    page_icon="📊",
    layout="wide"
)

st.title("📊 ○×試行管理表")


# =========================================================
# 定数
# =========================================================

ITEM_COL = "項目名"
TYPE_COL = "種類"
PERCENT_COL = "○割合"
MEMO = "メモ"


# =========================================================
# 初期データ作成
# =========================================================

def create_initial_data(trial_count=10, item_count=10):

    columns = (
        [ITEM_COL]
        + [str(i) for i in range(1, trial_count + 1)]
        + [PERCENT_COL]
    )

    rows = []

    for i in range(item_count):

        # 判定行
        choice_row = {
            ITEM_COL: f"項目{i + 1}",
            TYPE_COL: "判定",
            PERCENT_COL: 0.0
        }

        # メモ行
        memo_row = {
            ITEM_COL: MEMO,
            TYPE_COL: MEMO,
            PERCENT_COL: ""
        }

        for col in columns[1:-1]:
            choice_row[col] = ""
            memo_row[col] = ""

        rows.append(choice_row)
        rows.append(memo_row)

    df = pd.DataFrame(rows)

    return df


# =========================================================
# セッション状態
# =========================================================

if "df" not in st.session_state:

    st.session_state.df = create_initial_data()


# =========================================================
# 試行番号列を取得
# =========================================================

def get_trial_columns(df):

    return sorted(
        [
            str(col)
            for col in df.columns
            if str(col).isdigit()
        ],
        key=lambda x: int(x)
    )


# =========================================================
# ○割合計算
# =========================================================

def calculate_percentage(df):

    trial_columns = get_trial_columns(df)

    if not trial_columns:
        return df

    max_trial = max(
        int(col)
        for col in trial_columns
    )

    for row in range(0, len(df), 2):

        o_count = 0

        for col in trial_columns:

            if str(df.at[row, col]) == "○":

                o_count += 1

        if max_trial > 0:

            percentage = (
                o_count /
                max_trial
            ) * 100

        else:

            percentage = 0

        df.at[row, PERCENT_COL] = percentage

        if row + 1 < len(df):

            df.at[
                row + 1,
                PERCENT_COL
            ] = ""

    return df


# =========================================================
# 最後の列に入力されたら列追加
# =========================================================

def auto_add_column(df):

    trial_columns = get_trial_columns(df)

    if not trial_columns:
        df["1"] = ""
        return df

    last_column = trial_columns[-1]

    # 最後の列に何か入力されたか
    used = (
        df[last_column]
        .astype(str)
        .str.strip()
        .ne("")
        .any()
    )

    if used:

        new_column = str(
            int(last_column) + 1
        )

        df[new_column] = ""

    return df


# =========================================================
# 最後の項目に入力されたら行追加
# =========================================================

def auto_add_row(df):

    if len(df) < 2:
        return df

    last_choice_row = len(df) - 2

    trial_columns = get_trial_columns(df)

    used = False

    # 項目名
    item_name = str(
        df.at[
            last_choice_row,
            ITEM_COL
        ]
    )

    if (
        item_name.strip() != ""
        and item_name != f"項目{last_choice_row // 2 + 1}"
    ):

        used = True

    # ○×
    for col in trial_columns:

        if str(
            df.at[
                last_choice_row,
                col
            ]
        ).strip() != "":

            used = True
            break

    # メモ
    for col in trial_columns:

        if str(
            df.at[
                last_choice_row + 1,
                col
            ]
        ).strip() != "":

            used = True
            break

    if used:

        new_number = (
            len(df) // 2 + 1
        )

        choice_row = {
            ITEM_COL: f"項目{new_number}",
            TYPE_COL: "判定",
            PERCENT_COL: 0.0
        }

        memo_row = {
            ITEM_COL: MEMO,
            TYPE_COL: MEMO,
            PERCENT_COL: ""
        }

        for col in trial_columns:

            choice_row[col] = ""
            memo_row[col] = ""

        df = pd.concat(
            [
                df,
                pd.DataFrame(
                    [
                        choice_row,
                        memo_row
                    ]
                )
            ],
            ignore_index=True
        )

    return df


# =========================================================
# 表の列順整理
# =========================================================

def organize_columns(df):

    trial_columns = get_trial_columns(df)

    columns = (
        [ITEM_COL, TYPE_COL]
        + trial_columns
        + [PERCENT_COL]
    )

    for col in columns:

        if col not in df.columns:

            df[col] = ""

    return df[columns]


# =========================================================
# ○×の色設定
# =========================================================

def color_choice(value):

    if value == "○":

        return (
            "color: red; "
            "font-weight: bold; "
            "text-align: center;"
        )

    if value == "×":

        return (
            "color: blue; "
            "font-weight: bold; "
            "text-align: center;"
        )

    return "text-align: center;"


# =========================================================
# サイドバー
# =========================================================

st.sidebar.header("操作")


# ---------------------------------------------------------
# 行追加
# ---------------------------------------------------------

if st.sidebar.button("➕ 行を追加"):

    df = st.session_state.df.copy()

    trial_columns = get_trial_columns(df)

    new_number = (
        len(df) // 2 + 1
    )

    choice_row = {
        ITEM_COL: f"項目{new_number}",
        TYPE_COL: "判定",
        PERCENT_COL: 0.0
    }

    memo_row = {
        ITEM_COL: MEMO,
        TYPE_COL: MEMO,
        PERCENT_COL: ""
    }

    for col in trial_columns:

        choice_row[col] = ""
        memo_row[col] = ""

    df = pd.concat(
        [
            df,
            pd.DataFrame(
                [
                    choice_row,
                    memo_row
                ]
            )
        ],
        ignore_index=True
    )

    st.session_state.df = df

    st.rerun()


# ---------------------------------------------------------
# 列追加
# ---------------------------------------------------------

if st.sidebar.button("➕ 列を追加"):

    df = st.session_state.df.copy()

    trial_columns = get_trial_columns(df)

    if trial_columns:

        new_column = str(
            max(
                int(col)
                for col in trial_columns
            ) + 1
        )

    else:

        new_column = "1"

    df[new_column] = ""

    st.session_state.df = (
        organize_columns(df)
    )

    st.rerun()


# ---------------------------------------------------------
# 新規作成
# ---------------------------------------------------------

if st.sidebar.button("🔄 新規作成"):

    st.session_state.df = (
        create_initial_data()
    )

    st.rerun()


# =========================================================
# Excel読み込み
# =========================================================

st.sidebar.subheader("ファイル")


uploaded_file = st.sidebar.file_uploader(
    "Excelを読み込む",
    type=["xlsx"]
)


if uploaded_file is not None:

    try:

        workbook = load_workbook(
            uploaded_file,
            data_only=True
        )

        worksheet = workbook.active

        max_row = worksheet.max_row
        max_column = worksheet.max_column

        # 1行目から試行回数取得
        trial_columns = []

        for col in range(
            2,
            max_column
        ):

            value = worksheet.cell(
                row=1,
                column=col
            ).value

            if value is not None:

                trial_columns.append(
                    str(value)
                )

        rows = []

        for row in range(
            2,
            max_row + 1
        ):

            data = {
                ITEM_COL:
                    worksheet.cell(
                        row=row,
                        column=1
                    ).value
            }

            for index, trial in enumerate(
                trial_columns,
                start=2
            ):

                data[trial] = (
                    worksheet.cell(
                        row=row,
                        column=index
                    ).value
                    or ""
                )

            rows.append(data)

        imported_df = pd.DataFrame(rows)

        # 判定 / メモを設定
        imported_df[TYPE_COL] = ""

        for i in range(
            len(imported_df)
        ):

            if i % 2 == 0:

                imported_df.at[
                    i,
                    TYPE_COL
                ] = "判定"

            else:

                imported_df.at[
                    i,
                    TYPE_COL
                ] = MEMO

        if PERCENT_COL not in imported_df.columns:

            imported_df[
                PERCENT_COL
            ] = ""

        imported_df = organize_columns(
            imported_df
        )

        st.session_state.df = (
            imported_df
        )

    except Exception as e:

        st.sidebar.error(
            f"読み込みエラー: {e}"
        )


# =========================================================
# データ編集
# =========================================================

df = st.session_state.df.copy()

df = organize_columns(df)

df = auto_add_row(df)

df = auto_add_column(df)

df = calculate_percentage(df)


# =========================================================
# 編集用DataFrame
# =========================================================

trial_columns = get_trial_columns(df)


# ○×選択肢を作る
choice_columns = {}

for col in trial_columns:

    choice_columns[col] = st.column_config.SelectboxColumn(
        str(col),
        options=[
            "",
            "○",
            "×"
        ],
        width="small"
    )


# 項目名
column_config = {

    ITEM_COL:
        st.column_config.TextColumn(
            "項目名 / メモ",
            width="medium"
        ),

    TYPE_COL:
        st.column_config.TextColumn(
            "種類",
            disabled=True,
            width="small"
        ),

    PERCENT_COL:
        st.column_config.NumberColumn(
            "○割合",
            format="%.1f%%",
            disabled=True,
            width="small"
        )
}

column_config.update(
    choice_columns
)


# =========================================================
# 表
# =========================================================

st.subheader(
    "📊 表計算エリア"
)


edited_df = st.data_editor(

    df,

    column_config=column_config,

    use_container_width=True,

    num_rows="dynamic",

    hide_index=True,

    key="spreadsheet"
)


# =========================================================
# 編集結果反映
# =========================================================

# 編集されたDataFrameを保存
edited_df = edited_df.copy()

# 判定 / メモ
for row in range(
    len(edited_df)
):

    if row % 2 == 0:

        edited_df.at[
            row,
            TYPE_COL
        ] = "判定"

    else:

        edited_df.at[
            row,
            TYPE_COL
        ] = MEMO

        edited_df.at[
            row,
            ITEM_COL
        ] = MEMO


# 列整理
edited_df = organize_columns(
    edited_df
)

# 自動追加
edited_df = auto_add_row(
    edited_df
)

edited_df = auto_add_column(
    edited_df
)

# 割合計算
edited_df = calculate_percentage(
    edited_df
)

st.session_state.df = edited_df


# =========================================================
# ○ / × 色表示用の表
# =========================================================

st.subheader(
    "🎨 入力結果"
)


display_df = edited_df.copy()


def style_table(row):

    styles = []

    for column in row.index:

        value = row[column]

        if value == "○":

            styles.append(
                "color: red; "
                "font-weight: bold; "
                "text-align: center;"
            )

        elif value == "×":

            styles.append(
                "color: blue; "
                "font-weight: bold; "
                "text-align: center;"
            )

        else:

            styles.append("")

    return styles


styled_df = display_df.style.apply(
    style_table,
    axis=1
)

st.dataframe(
    styled_df,
    use_container_width=True,
    hide_index=True
)


# =========================================================
# 集計
# =========================================================

st.subheader(
    "📈 ○割合"
)


trial_columns = get_trial_columns(
    st.session_state.df
)

max_trial = max(
    [
        int(col)
        for col in trial_columns
    ],
    default=0
)


summary = []


for row in range(
    0,
    len(st.session_state.df),
    2
):

    item_name = st.session_state.df.at[
        row,
        ITEM_COL
    ]

    o_count = 0

    for col in trial_columns:

        if (
            st.session_state.df.at[
                row,
                col
            ] == "○"
        ):

            o_count += 1

    if max_trial > 0:

        percentage = (
            o_count /
            max_trial
        ) * 100

    else:

        percentage = 0

    summary.append(
        {
            "項目名":
                item_name,

            "○回数":
                o_count,

            "全体試行回数":
                max_trial,

            "○割合":
                f"{percentage:.1f}%"
        }
    )


summary_df = pd.DataFrame(
    summary
)


st.dataframe(
    summary_df,
    use_container_width=True,
    hide_index=True
)


# =========================================================
# Excel保存
# =========================================================

st.subheader(
    "💾 保存"
)


excel_buffer = io.BytesIO()


workbook = Workbook()

worksheet = workbook.active

worksheet.title = (
    "試行管理表"
)


# A1は空白
worksheet.cell(
    row=1,
    column=1,
    value=""
)


# 1行目
for index, col in enumerate(
    trial_columns,
    start=2
):

    worksheet.cell(
        row=1,
        column=index,
        value=int(col)
    )


# ○割合
worksheet.cell(
    row=1,
    column=len(trial_columns) + 2,
    value="○割合"
)


# データ
for row_index in range(
    len(st.session_state.df)
):

    excel_row = row_index + 2

    worksheet.cell(
        row=excel_row,
        column=1,
        value=
        st.session_state.df.at[
            row_index,
            ITEM_COL
        ]
    )

    for col_index, col in enumerate(
        trial_columns,
        start=2
    ):

        value = (
            st.session_state.df.at[
                row_index,
                col
            ]
        )

        cell = worksheet.cell(
            row=excel_row,
            column=col_index,
            value=value
        )

        # ○ = 赤
        if value == "○":

            cell.font = cell.font.copy(
                color="FF0000"
            )

        # × = 青
        elif value == "×":

            cell.font = cell.font.copy(
                color="0000FF"
            )

    # 割合
    worksheet.cell(
        row=excel_row,
        column=len(trial_columns) + 2,
        value=
        st.session_state.df.at[
            row_index,
            PERCENT_COL
        ]
    )


# 列幅
worksheet.column_dimensions[
    "A"
].width = 20

for col in range(
    2,
    len(trial_columns) + 2
):

    worksheet.column_dimensions[
        worksheet.cell(
            row=1,
            column=col
        ).column_letter
    ].width = 12


worksheet.column_dimensions[
    worksheet.cell(
        row=1,
        column=len(trial_columns) + 2
    ).column_letter
].width = 12


workbook.save(
    excel_buffer
)


excel_buffer.seek(0)


st.download_button(

    label="📊 Excelとして保存",

    data=excel_buffer.getvalue(),

    file_name="○×試行管理表.xlsx",

    mime=(
        "application/"
        "vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet"
    )
)


# =========================================================
# CSV保存
# =========================================================

csv_data = (
    st.session_state.df
    .to_csv(
        index=False,
        encoding="utf-8-sig"
    )
)


st.download_button(

    label="📥 CSVとして保存",

    data=csv_data,

    file_name="○×試行管理表.csv",

    mime="text/csv"
)


# =========================================================
# 使い方
# =========================================================

with st.expander(
    "📖 使い方"
):

    st.write(
        """
1. 項目名を入力します。
2. 各試行セルをクリックすると「○」「×」「空欄」から選択できます。
3. 判定セルの下がメモ欄です。
4. ○は赤、×は青で表示されます。
5. 最後の試行列に入力すると、新しい列が自動追加されます。
6. 最後の項目に入力すると、新しい行が自動追加されます。
7. ○割合は「1行目の数字の最大値」を全体試行回数として計算します。
8. Excel保存したデータは左側から読み込めます。
9. Streamlit Cloudなどに公開すればURLでアクセスできます。
        """
    )
